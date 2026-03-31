"""Excel export utility for survey responses."""

import json
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# Columns that contain JSON-encoded lists
JSON_COLUMNS = [
    "subjects_taught",
    "grade_levels",
    "ai_tools_used",
    "innovative_uses",
    "barriers",
    "support_needed",
]


def _json_to_string(val):
    """Convert a JSON-encoded list to a comma-separated string."""
    if pd.isna(val) or val == "":
        return ""
    try:
        items = json.loads(val)
        if isinstance(items, list):
            return ", ".join(items)
    except (json.JSONDecodeError, TypeError):
        pass
    return str(val)


def create_excel_download(df: pd.DataFrame) -> bytes:
    """Create a formatted Excel file with responses and summary sheets."""
    wb = Workbook()

    # --- Sheet 1: All Responses ---
    ws_responses = wb.active
    ws_responses.title = "Responses"

    # Prepare display dataframe
    display_df = df.copy()
    if "id" in display_df.columns:
        display_df = display_df.drop(columns=["id"])
    for col in JSON_COLUMNS:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(_json_to_string)

    # Column name mapping for readability
    col_labels = {
        "submitted_at": "Submitted At",
        "school_board": "School Board",
        "subjects_taught": "Subjects Taught",
        "grade_levels": "Grade Levels",
        "experience_years": "Years of Experience",
        "city_tier": "City Tier",
        "ai_tools_used": "AI Tools Used",
        "discovery_channel": "How Discovered AI",
        "ai_usage_duration": "AI Usage Duration",
        "freq_lesson_plans": "Freq: Lesson Plans",
        "freq_assessments": "Freq: Assessments",
        "freq_personalized": "Freq: Personalized Materials",
        "freq_content": "Freq: Classroom Content",
        "freq_admin": "Freq: Admin Tasks",
        "freq_engagement": "Freq: Engagement Activities",
        "freq_grading": "Freq: Grading",
        "freq_parent_comm": "Freq: Parent Communication",
        "innovative_uses": "Innovative Uses",
        "innovative_desc": "Innovative Use Description",
        "impact_learning": "Impact: Learning Outcomes",
        "impact_engagement": "Impact: Student Engagement",
        "impact_individual": "Impact: Individual Weaknesses",
        "impact_performance": "Impact: Exam Performance",
        "impact_creativity": "Impact: Creativity",
        "barriers": "Barriers",
        "support_needed": "Support Needed",
        "future_likelihood": "Future AI Use Likelihood",
        "future_priority": "Future Priority Area",
    }

    header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # Write headers
    headers = [col_labels.get(c, c) for c in display_df.columns]
    for col_idx, header in enumerate(headers, 1):
        cell = ws_responses.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_idx, (_, row) in enumerate(display_df.iterrows(), 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws_responses.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap_alignment

    # Auto-size columns (approximate)
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row_idx in range(2, min(len(display_df) + 2, 52)):  # Sample first 50 rows
            cell_val = ws_responses.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_length = max(max_length, min(len(str(cell_val)), 50))
        ws_responses.column_dimensions[get_column_letter(col_idx)].width = max_length + 4

    # Freeze top row
    ws_responses.freeze_panes = "A2"

    # --- Sheet 2: Summary Statistics ---
    ws_summary = wb.create_sheet("Summary")
    ws_summary.cell(row=1, column=1, value="Question").font = Font(bold=True, size=11)
    ws_summary.cell(row=1, column=2, value="Option").font = Font(bold=True, size=11)
    ws_summary.cell(row=1, column=3, value="Count").font = Font(bold=True, size=11)
    ws_summary.cell(row=1, column=4, value="Percentage").font = Font(bold=True, size=11)

    for col_idx in range(1, 5):
        ws_summary.cell(row=1, column=col_idx).fill = header_fill
        ws_summary.cell(row=1, column=col_idx).font = header_font

    summary_row = 2
    total_responses = len(df)

    # Single-choice columns
    single_cols = [
        "school_board", "experience_years", "city_tier",
        "discovery_channel", "ai_usage_duration",
        "freq_lesson_plans", "freq_assessments", "freq_personalized",
        "freq_content", "freq_admin", "freq_engagement",
        "freq_grading", "freq_parent_comm",
        "impact_learning", "impact_engagement", "impact_individual",
        "impact_performance", "impact_creativity",
        "future_likelihood", "future_priority",
    ]

    for col in single_cols:
        if col not in df.columns:
            continue
        label = col_labels.get(col, col)
        counts = df[col].dropna().value_counts()
        for option, count in counts.items():
            pct = f"{count / total_responses * 100:.1f}%" if total_responses > 0 else "0%"
            ws_summary.cell(row=summary_row, column=1, value=label)
            ws_summary.cell(row=summary_row, column=2, value=option)
            ws_summary.cell(row=summary_row, column=3, value=count)
            ws_summary.cell(row=summary_row, column=4, value=pct)
            summary_row += 1
        summary_row += 1  # Blank row between questions

    # Multi-choice columns
    for col in JSON_COLUMNS:
        if col not in df.columns:
            continue
        label = col_labels.get(col, col)
        all_values = []
        for val in df[col].dropna():
            try:
                items = json.loads(val)
                if isinstance(items, list):
                    all_values.extend(items)
            except (json.JSONDecodeError, TypeError):
                if val:
                    all_values.append(val)
        if all_values:
            counts = pd.Series(all_values).value_counts()
            for option, count in counts.items():
                pct = f"{count / total_responses * 100:.1f}%" if total_responses > 0 else "0%"
                ws_summary.cell(row=summary_row, column=1, value=label)
                ws_summary.cell(row=summary_row, column=2, value=option)
                ws_summary.cell(row=summary_row, column=3, value=count)
                ws_summary.cell(row=summary_row, column=4, value=pct)
                summary_row += 1
            summary_row += 1

    # Auto-size summary columns
    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 55
    ws_summary.column_dimensions["C"].width = 10
    ws_summary.column_dimensions["D"].width = 12
    ws_summary.freeze_panes = "A2"

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
